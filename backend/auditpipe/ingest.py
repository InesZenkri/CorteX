"""Ingestion. Structured German files (GDPdU .txt, .csv) are parsed
deterministically with cell-level provenance — no LLM needed and no chance of a
hallucinated figure. Unstructured files (.pdf, .docx) go to GPT-5.6 for
classification + extraction under the verbatim-quote contract. .xlsx is read
deterministically via openpyxl.

Schemas for the eight GDPdU tables are declared here; unknown .csv files are
parsed header-driven. Everything lands as Row objects with provenance.
"""

from __future__ import annotations
import csv
import datetime as dt
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from . import config

# ---- GDPdU schema registry (column names; roles drive normalization) ------
# role: key|text|num|date|time|bool
GDPDU_SCHEMAS: dict[str, list[tuple[str, str]]] = {
    "Sachkonten.txt": [
        ("SACHKONTONUMMER", "key"), ("SACHKONTONAME", "text"), ("SACHKONTOTYP", "text"),
        ("SPERRE", "text"), ("EXCLUSIVBENUTZER", "text"), ("BENUTZUNG", "text"),
        ("KONTENART", "text")],
    "Sachkontobuchungen.txt": [
        ("SACHKONTONUMMER", "key"), ("PERIODENCODE", "text"), ("STEUERREF", "text"),
        ("PERIODENZUGEH", "text"), ("BUCHUNGSTYP", "text"), ("KORREKTUR", "bool"),
        ("HABENBUCHUNG", "bool"), ("BUCHUNGSBETRAG", "num"), ("WAEHRUNG", "text"),
        ("BUCHUNGSWERT", "num"), ("BUCHUNGSTEXT", "text"), ("BUCHUNGSDATUM", "date"),
        ("BUCHUNGSNUMMER", "key"), ("BELEGDATUM", "date"), ("BELEGNUMMER", "key"),
        ("SPEZIALBUCHUNG", "text"), ("ERFASSUNGSNUMMER", "text"), ("JOURNALZEILE", "text"),
        ("GEGENKONTO", "key"), ("DOKUMENT", "text"), ("BENUTZERKENNUNG", "key"),
        ("ERFASSUNGSDATUM", "date"), ("ERFASSUNGSZEIT", "time"), ("FESTSCHREIBUNG", "bool")],
    "Kunden.txt": [
        ("KUNDENNUMMER", "key"), ("USTID", "key"), ("STRASSE", "text"), ("PLZ", "text"),
        ("ORT", "text"), ("LAND", "text"), ("NAME", "text"), ("KONTENGRUPPE", "text"),
        ("F9", "text"), ("F10", "text"), ("F11", "text"), ("KLASSE", "text"), ("WAEHRUNG", "text")],
    "Lieferanten.txt": [
        ("LIEFERANTENNUMMER", "key"), ("USTID", "key"), ("STRASSE", "text"), ("PLZ", "text"),
        ("ORT", "text"), ("LAND", "text"), ("NAME", "text"), ("KONTENGRUPPE", "text"),
        ("F9", "text"), ("KLASSE", "text"), ("WAEHRUNG", "text")],
    "Anlagen.txt": [
        ("ANLAGENNUMMER", "key"), ("ANLAGENNAME", "text"), ("SACHKONTONUMMER", "key"),
        ("ANLAGENTYP", "text"), ("F5", "text"), ("F6", "text"), ("F7", "text"),
        ("F8", "text"), ("STATUS", "text")],
    "Kundenbuchungen.txt": [
        ("KUNDENNUMMER", "key"), ("BELEGNUMMER", "key"), ("BUCHUNGSDATUM", "date"),
        ("F4", "text"), ("BELEGDATUM", "date"), ("BUCHUNGSTEXT", "text"),
        ("BUCHUNGSBETRAG", "num"), ("WAEHRUNG", "text"), ("BUCHUNGSWERT", "num"),
        ("F10", "text"), ("F11", "text"), ("BUCHUNGSTYP", "text")],
    "Lieferantenbuchungen.txt": [
        ("LIEFERANTENNUMMER", "key"), ("BELEGNUMMER", "key"), ("BUCHUNGSDATUM", "date"),
        ("F4", "text"), ("BELEGDATUM", "date"), ("BUCHUNGSTEXT", "text"),
        ("BUCHUNGSBETRAG", "num"), ("WAEHRUNG", "text"), ("BUCHUNGSWERT", "num"),
        ("F10", "text"), ("F11", "text"), ("BUCHUNGSTYP", "text"), ("HABENBUCHUNG", "bool")],
    "Anlagenbuchungen.txt": [
        ("ANLAGENNUMMER", "key"), ("BUCHUNGSDATUM", "date"), ("BELEGNUMMER", "key"),
        ("BUCHUNGSBETRAG", "num"), ("WAEHRUNG", "text"), ("BUCHUNGSWERT", "num"),
        ("BUCHUNGSTYP", "text"), ("F8", "text"), ("SACHKONTONUMMER", "key"), ("STATUS", "text")],
}


@dataclass
class Cell:
    column: str
    raw: str
    value: Any
    role: str


@dataclass
class Row:
    table: str
    source_file: str
    line_no: int
    cells: dict[str, Cell] = field(default_factory=dict)

    def val(self, c): return self.cells[c].value if c in self.cells else None
    def raw(self, c): return self.cells[c].raw if c in self.cells else ""


# ---- normalizers ----------------------------------------------------------
def _num(raw: str):
    s = raw.strip()
    if not s:
        return None
    neg = s.startswith("-")
    s = s.lstrip("+-").replace(".", "").replace(",", ".")
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return -d if neg else d


def _date(raw: str):
    s = raw.strip()
    if not s:
        return None
    try:
        return dt.datetime.strptime(s, config.DATE_FMT).date().isoformat()
    except ValueError:
        return None


def _time(raw: str):
    s = raw.strip()
    try:
        return dt.datetime.strptime(s, "%H:%M:%S").time().isoformat()
    except ValueError:
        return None


def _bool(raw: str):
    s = raw.strip().lower()
    if s in ("ja", "yes", "j", "y", "true", "1", "-1"):
        return True
    if s in ("nein", "no", "n", "false", "0"):
        return False
    return None


def _norm(role: str, raw: str):
    return {"num": _num, "date": _date, "time": _time, "bool": _bool}.get(
        role, lambda x: x.strip())(raw)


# ---- structured parsers ---------------------------------------------------
def parse_gdpdu(path: Path) -> list[Row]:
    schema = GDPDU_SCHEMAS[path.name]
    table = path.stem
    rows: list[Row] = []
    with path.open(encoding=config.GDPDU_ENCODING, newline="") as fh:
        for i, line in enumerate(fh, 1):
            line = line.rstrip("\r\n")
            if not line:
                continue
            fields = next(csv.reader([line], delimiter=config.CSV_DELIMITER, quotechar=config.TEXT_QUOTE))
            fields = (fields + [""] * len(schema))[:len(schema)]
            r = Row(table, path.name, i)
            for (col, role), raw in zip(schema, fields):
                r.cells[col] = Cell(col, raw, _norm(role, raw), role)
            rows.append(r)
    return rows


def _guess_role(col: str) -> str:
    c = col.upper()
    if any(k in c for k in ("BETRAG", "SUMME", "SALDO", "LIMIT", "AUSNUTZUNG", "PROZENT", "ANZAHL", "_EUR")):
        return "num"
    if any(k in c for k in ("DATUM", "_AM")):
        return "date"
    return "text"


def parse_csv(path: Path) -> list[Row]:
    table = path.stem.replace("-", "_")
    # sniff encoding: try utf-8 then cp1252
    for enc in ("utf-8", "cp1252", "ISO-8859-1"):
        try:
            with path.open(encoding=enc, newline="") as fh:
                raw_rows = list(csv.reader(fh, delimiter=config.CSV_DELIMITER, quotechar=config.TEXT_QUOTE))
            break
        except UnicodeDecodeError:
            continue
    header = [h.strip() for h in raw_rows[0]]
    roles = [_guess_role(h) for h in header]
    rows: list[Row] = []
    for i, rec in enumerate(raw_rows[1:], 2):
        if not any(x.strip() for x in rec):
            continue
        rec = (rec + [""] * len(header))[:len(header)]
        r = Row(table, path.name, i)
        for col, role, raw in zip(header, roles, rec):
            r.cells[col] = Cell(col, raw, _norm(role, raw), role)
        rows.append(r)
    return rows


def read_xlsx_text(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"# sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            out.append("\t".join("" if c is None else str(c) for c in row))
    return "\n".join(out)


def parse_xlsx_structured(path: Path) -> list[Row]:
    """Load each sheet as a generic table. Header = first row with >=3 non-empty
    string cells (skips title banners). Enables structural detectors (e.g. the
    authorization matrix) to work without the LLM."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    rows_out: list[Row] = []
    for ws in wb.worksheets:
        data = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, r in enumerate(data):
            noncell = [c for c in r if c is not None and str(c).strip()]
            if len(noncell) >= 3:
                header_idx = i
                break
        if header_idx is None:
            continue
        header = [str(c).strip() if c is not None else f"C{j}"
                  for j, c in enumerate(data[header_idx])]
        header = [h or f"C{j}" for j, h in enumerate(header)]
        table = f"{path.stem}_{ws.title}".replace("-", "_").replace(" ", "_")
        for li, rec in enumerate(data[header_idx + 1:], header_idx + 2):
            if not any(c is not None and str(c).strip() for c in rec):
                continue
            r = Row(table, path.name, li)
            for col, val in zip(header, rec):
                raw = "" if val is None else str(val)
                r.cells[col] = Cell(col, raw, raw.strip(), "text")
            rows_out.append(r)
    return rows_out


def read_pdf_text(path: Path) -> str:
    try:
        from pypdf import PdfReader
        return "\n".join((p.extract_text() or "") for p in PdfReader(str(path)).pages)
    except Exception:
        return ""


def read_docx_text(path: Path) -> str:
    try:
        import docx  # python-docx
        d = docx.Document(str(path))
        return "\n".join(p.text for p in d.paragraphs)
    except Exception:
        return ""


# ---- dispatcher -----------------------------------------------------------
STRUCTURED_SUFFIXES = {".txt", ".csv"}
UNSTRUCTURED_SUFFIXES = {".pdf", ".docx"}


def discover(data_dir: str) -> list[Path]:
    """Recursively find ingestible files. Dossiers ship nested (e.g.
    data/Sachkonten/*.txt, data/Begleitdokumente/*.csv) — a flat scan misses them."""
    skip = {"index.xml", ".ds_store"}
    allowed = STRUCTURED_SUFFIXES | UNSTRUCTURED_SUFFIXES | {".xlsx"}
    return sorted(
        p for p in Path(data_dir).rglob("*")
        if p.is_file()
        and p.suffix.lower() in allowed
        and p.name.lower() not in skip
    )
